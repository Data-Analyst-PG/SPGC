import streamlit as st
import pandas as pd
from io import BytesIO
from supabase import create_client, Client
from datetime import date
from streamlit.runtime.scriptrunner import StopException
import numpy as np
import re
import openpyxl

# --- CONFIGURACIÓN SUPABASE ---
url = st.secrets["SUPABASE_URL"]
key = st.secrets["SUPABASE_KEY"]
supabase = create_client(url, key)



# ======================================================================================================
st.title("🧾Prorrateo de Gastos Generales")

# ================================
# Subir archivo y generar resumen
# ================================

# Subir archivo Excel
uploaded_file = st.file_uploader("Sube el archivo con la hoja 'PASO 1'", type=["xlsx"])

if uploaded_file:
    try:
        # Leer hoja específica
        df = pd.read_excel(uploaded_file, sheet_name="PASO 1")
        df.columns = df.columns.str.strip().str.upper()

        required = {"SUCURSAL", "AREA/CUENTA", "CARGOS"}
        missing = required - set(df.columns)
        if missing:
            st.error(f"❌ Faltan columnas requeridas en PASO 1: {sorted(missing)}")
            st.stop()

        st.session_state["df_original"] = df

        # Filtrar COMUNES (antes "GASTO GENERAL"): INTERNO + EXTERNO
        df["SUCURSAL"] = df["SUCURSAL"].astype(str).str.strip().str.upper()
        comunes_base = df[df["SUCURSAL"].isin(["GASTO GENERAL", "INTERNO", "EXTERNO"])].copy()

        # Validación clara
        if comunes_base.empty:
            st.error("❌ No hay filas con SUCURSAL = GASTO GENERAL / INTERNO / EXTERNO. No hay comunes para resumir.")
            st.stop()

        # Agrupar por AREA/CUENTA y sumar los CARGOS
        resumen = (
            comunes_base
            .groupby("AREA/CUENTA", as_index=False)["CARGOS"]
            .sum()
            .sort_values(by="CARGOS", ascending=False)
        )

        # Guardar en session_state para el módulo 2
        st.session_state['resumen'] = resumen

        st.success("Resumen generado con éxito.")
        st.dataframe(resumen, use_container_width=True)

        # Exportar a Excel
        def export_excel(df):
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
                df.to_excel(writer, sheet_name="Resumen Gastos", index=False)
            return buffer.getvalue()

        st.download_button(
            "📥 Descargar resumen en Excel",
            data=export_excel(resumen),
            file_name="resumen_gastos_generales.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except StopException:
        raise
    except Exception as e:
        st.error(f"Error procesando el archivo: {e}")

# ========================================
# Captura de Tráfico y Fecha por Sucursal
# ========================================

st.subheader("🚛 Captura de Tráfico por Sucursal")

# Lista fija de sucursales (precargadas)
sucursales = [
    "CAR-GAR", "CHICAGO", "CONSOLIDADO", "DALLAS", "GUADALAJARA",
    "LEON", "LINCOLN LOGISTICS", "MG HAULERS", "MONTERREY",
    "NUEVO LAREDO", "QUERETARO", "ROLANDO ALFARO"
]

# Fecha global que se aplicará a todos los registros
fecha_global = st.date_input("📅 Fecha de tráfico", value=date.today())

# Crear un DataFrame editable para capturar los números de tráfico
traficos_df = pd.DataFrame({
    "Sucursal": sucursales,
    "Tráfico": ["" for _ in sucursales]
})

st.markdown("### ✏️ Captura los números de tráfico")
edit_df = st.data_editor(
    traficos_df,
    use_container_width=True,
    num_rows="fixed",
    key="trafico_editor"
)

# Botón para guardar en Supabase
if st.button("💾 Guardar en Supabase", key="save_traficos"):
    try:
        df_to_save = edit_df.copy()

        # Validaciones básicas
        faltan = df_to_save["Tráfico"].astype(str).str.strip().eq("")
        if faltan.any():
            st.warning("Completa todos los números de tráfico antes de guardar.")
        else:
            # Normaliza texto y FECHA -> string ISO (YYYY-MM-DD)
            df_to_save["Sucursal"] = df_to_save["Sucursal"].astype(str).str.strip()
            df_to_save["Tráfico"]  = df_to_save["Tráfico"].astype(str).str.strip()
            df_to_save["Fecha"]    = pd.to_datetime(fecha_global).strftime("%Y-%m-%d")

            # Renombra a las columnas reales de la tabla
            payload = (
                df_to_save[["Sucursal", "Tráfico", "Fecha"]]
                .rename(columns={"Tráfico": "Trafico"})
                .to_dict(orient="records")
            )

            # Inserta (o usa upsert si quieres evitar duplicados Sucursal+Fecha)
            res = supabase.table("viajes_distribucion").insert(payload).execute()
            # Si prefieres actualizar/enlazar:
            # res = supabase.table("viajes_distribucion") \
            #     .upsert(payload, on_conflict="Sucursal,Fecha").execute()

            st.success(f"✅ Tráficos guardados ({len(payload)} filas).")
    except Exception as e:
        st.error(f"Error al guardar en Supabase: {e}")

st.divider()

# ======================================================================================================
st.title("📘Catálogo de Distribución por AREA/CUENTA")

tipos_distribucion = [
    "Facturación Dlls", "MC", "Tráficos",
    "Empleado hub", "Empleados mv", "XTRALEASE", "Uso Cajas"
]

# Validamos que venga del Módulo 1
if 'resumen' in st.session_state:
    resumen = st.session_state['resumen']
    resumen = resumen[['AREA/CUENTA']].drop_duplicates().reset_index(drop=True)

    # Cargar catálogo existente desde Supabase
    data_supabase = supabase.table("catalogo_distribucion").select("*").execute().data
    catalogo_existente = pd.DataFrame(data_supabase)

    # Unir resumen con catálogo existente
    if not catalogo_existente.empty:
        catalogo_existente = catalogo_existente.rename(columns={
            "area_cuenta": "AREA/CUENTA",
            "tipo_distribucion": "TIPO DISTRIBUCIÓN"
        })
        resumen_merged = resumen.merge(catalogo_existente, on="AREA/CUENTA", how="left")
    else:
        resumen_merged = resumen.copy()
        resumen_merged["TIPO DISTRIBUCIÓN"] = None

    st.subheader("Catálogo de Distribución")
    resumen_merged = resumen_merged.sort_values(by=["TIPO DISTRIBUCIÓN", "AREA/CUENTA"], na_position="first").reset_index(drop=True)
    edited_df = st.data_editor(
        resumen_merged,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "TIPO DISTRIBUCIÓN": st.column_config.SelectboxColumn(
                label="Tipo de Distribución",
                options=tipos_distribucion,
                required=True
            )
        }
    )

    # Botón para guardar nuevos registros y actualizaciones en Supabase
    if st.button("💾 Guardar en Supabase", key="save_catalogo"):
        nuevos = edited_df[edited_df["TIPO DISTRIBUCIÓN"].notna()]
        for _, row in nuevos.iterrows():
            supabase.table("catalogo_distribucion").upsert({
                "area_cuenta": row["AREA/CUENTA"],
                "tipo_distribucion": row["TIPO DISTRIBUCIÓN"]
            }).execute()
        st.success("Catálogo actualizado en Supabase.")

else:
    st.warning("Primero genera el resumen de COMUNES (GASTO GENERAL / INTERNO / EXTERNO) en el Módulo 1.")

# ======================================================================================================
st.title("📊Datos Generales y Cálculo de Porcentajes")

# Subida del archivo con la hoja GTS
archivo_gts = st.file_uploader("📤 Sube el archivo con la hoja 'GTS'", type=["xlsx"])

if archivo_gts:
    try:
        df_gts = pd.read_excel(archivo_gts, sheet_name="GTS")

        # Limpiar encabezados
        df_gts.columns = df_gts.columns.str.strip().str.upper()

        # Mostrar tabla original
        st.subheader("📄 Datos GTS cargados:")
        st.dataframe(df_gts, use_container_width=True)

        # Seleccionar solo columnas numéricas (los tipos de distribución)
        tipo_distrib_cols = df_gts.columns.drop("SUCURSAL")

        # Calcular totales por tipo
        totales = df_gts[tipo_distrib_cols].sum().rename("TOTAL").to_frame().T

        st.subheader("📌 Totales por Tipo de Distribución")
        st.dataframe(totales, use_container_width=True)

        # Calcular porcentajes por sucursal y tipo
        porcentajes = df_gts.copy()
        for col in tipo_distrib_cols:
            total = df_gts[col].sum()
            if total != 0:
                porcentajes[col] = df_gts[col] / total
            else:
                porcentajes[col] = 0

        st.subheader("📈 Porcentaje de Participación")
        st.dataframe(porcentajes, use_container_width=True)

        # Guardar en memoria para módulo 4
        st.session_state["porcentajes"] = porcentajes.assign(
            SUCURSAL=lambda d: d["SUCURSAL"].astype(str).str.strip().str.upper()
        ).set_index("SUCURSAL")


    except Exception as e:
        st.error(f"Ocurrió un error al procesar la hoja GTS: {e}")

# ======================================================================================================
st.title("🔄Gasto General + Costos por Sucursal (con Tráfico/Fecha)")

# ---------- Comprobación de insumos previos ----------
if not all(k in st.session_state for k in ["df_original", "resumen", "porcentajes"]):
    st.warning("Faltan datos. Completa primero los módulos previos (df_original, resumen, porcentajes).")
    st.stop()

df_original = st.session_state["df_original"].copy()
resumen_gasto_general = st.session_state["resumen"].copy()  # agrupado por AREA/CUENTA
porcentajes = st.session_state["porcentajes"].copy()

# Normalizaciones
df_original.columns = df_original.columns.str.strip().str.upper()
porcentajes.columns = [str(c).upper() for c in porcentajes.columns]
if "CONCEPTO" not in df_original.columns:
    df_original["CONCEPTO"] = ""

# Catálogo (AREA/CUENTA -> TIPO DISTRIBUCIÓN)
cat_resp = supabase.table("catalogo_distribucion").select("*").execute()
catalogo = pd.DataFrame(cat_resp.data)
if catalogo.empty:
    st.error("No hay datos en 'catalogo_distribucion'.")
    st.stop()
catalogo = catalogo.rename(columns={
    "area_cuenta": "AREA/CUENTA",
    "tipo_distribucion": "TIPO DISTRIBUCIÓN"
})

# Viajes/fechas (para seleccionar fecha específica)
viajes_resp = supabase.table("viajes_distribucion").select("*").execute()
viajes = pd.DataFrame(viajes_resp.data)
if viajes.empty:
    st.info("No hay registros en 'viajes_distribucion'. Trafico/Fecha quedarán vacíos.")
    viajes = pd.DataFrame(columns=["Sucursal", "Trafico", "Fecha"])
viajes = viajes.rename(columns={"Sucursal": "SUCURSAL", "Trafico": "TRAFICO", "Fecha": "FECHA"})
viajes["SUCURSAL"] = viajes["SUCURSAL"].astype(str).str.upper().str.strip()
viajes["FECHA"] = pd.to_datetime(viajes["FECHA"], errors="coerce")

# Selector de fecha específica (usa las disponibles en la tabla)
fechas_disponibles = sorted(viajes["FECHA"].dropna().dt.date.unique())
fecha_elegida = st.date_input(
    "📅 Selecciona la FECHA de viajes_distribucion a usar",
    value=(fechas_disponibles[-1] if len(fechas_disponibles) else pd.Timestamp.today().date()),
    min_value=(fechas_disponibles[0] if len(fechas_disponibles) else pd.Timestamp.today().date())
)

viajes_sel = viajes[viajes["FECHA"].dt.date.eq(fecha_elegida)].copy()
viajes_sel["FECHA"] = viajes_sel["FECHA"].dt.strftime("%Y-%m-%d")  # ISO string
viajes_sel = viajes_sel.drop_duplicates(subset=["SUCURSAL"], keep="last")

# Helper para anexar Trafico/Fecha por sucursal
def anexar_trafico_fecha(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.assign(TRAFICO=None, FECHA=None)
    out = df.merge(viajes_sel[["SUCURSAL", "TRAFICO", "FECHA"]], on="SUCURSAL", how="left")
    return out

# ============ BLOQUE A: COSTOS CON SUCURSAL ASIGNADA ============
df_original["SUCURSAL"] = df_original["SUCURSAL"].astype(str).str.strip().str.upper()


def _suc_key(x: str) -> str:
    # Convierte "CAR GAR", "CAR-GAR", "CAR/GAR" -> "CARGAR"
    return re.sub(r"[^A-Z0-9]", "", str(x).strip().upper())

# Sucursales válidas basadas en GTS (canónicas)
if "porcentajes" in st.session_state:
    # porcentajes está indexado por SUCURSAL (ya normalizado en tu módulo 3)
    suc_validas = [s for s in st.session_state["porcentajes"].index.tolist()]
else:
    # fallback si por alguna razón no está porcentajes
    suc_validas = []

# Construir mapa key->nombre canónico (ej: "CARGAR" -> "CAR-GAR")
map_suc = {_suc_key(s): s for s in suc_validas}

# Homologar sucursal en df_original usando GTS como verdad
df_original["SUCURSAL_ORIG"] = df_original["SUCURSAL"]
df_original["SUCURSAL"] = df_original["SUCURSAL"].apply(lambda s: map_suc.get(_suc_key(s), s))

# (Opcional) avisar cuáles no se pudieron homologar (y no son comunes)
no_recon = df_original[
    (~df_original["SUCURSAL"].isin(["GASTO GENERAL", "INTERNO", "EXTERNO"])) &
    (~df_original["SUCURSAL"].isin(suc_validas))
]["SUCURSAL_ORIG"].dropna().astype(str).unique().tolist()

if no_recon:
    st.warning(f"Sucursales NO reconocidas (revisa ortografía): {no_recon[:15]}{'...' if len(no_recon)>15 else ''}")

# COSTO INDIRECTO = todo lo que NO sea INTERNO/EXTERNO
no_general = df_original[~df_original["SUCURSAL"].isin(["GASTO GENERAL", "INTERNO", "EXTERNO"])].copy()

# TIPO COSTO fijo para este bloque
no_general["TIPO COSTO"] = "COSTO INDIRECTO"
# TIPO DISTRIBUCIÓN fijo para este bloque
no_general["TIPO DISTRIBUCIÓN"] = "Costo fijo en sucursal"

# Agrupar por AREA/CUENTA + SUCURSAL + TIPO DISTRIBUCIÓN + TIPO COSTO
directos_agr = (
    no_general
    .assign(SUCURSAL=lambda d: d["SUCURSAL"].astype(str).str.upper().str.strip())
    .groupby(["AREA/CUENTA", "SUCURSAL", "TIPO DISTRIBUCIÓN", "TIPO COSTO"], as_index=False)["CARGOS"]
    .sum()
    .rename(columns={"CARGOS": "CARGO ASIGNADO"})
)

# Anexar Trafico/Fecha por sucursal (fecha seleccionada)
directos_agr = anexar_trafico_fecha(directos_agr)

# ============ BLOQUE B: GASTO GENERAL (Viejo) INTERNO + EXTERNO (Nuevo) (prorrateo) ============
# COMUNES = GASTO GENERAL (viejo) + INTERNO / EXTERNO (nuevo)
comunes = df_original[df_original["SUCURSAL"].isin(["GASTO GENERAL", "INTERNO", "EXTERNO"])].copy()

if comunes.empty:
    st.error("❌ No hay filas con SUCURSAL = GASTO GENERAL / INTERNO / EXTERNO para prorratear.")
    st.stop()

# SUCURSAL define el común
def tipo_costo_hibrido(row) -> str:
    suc = str(row.get("SUCURSAL", "")).strip().upper()

    # Nuevo esquema
    if suc == "INTERNO":
        return "COMUN INTERNO"
    if suc == "EXTERNO":
        return "COMUN EXTERNO"

    # Esquema viejo: GASTO GENERAL decide por CONCEPTO (IN/EX)
    if suc == "GASTO GENERAL":
        c = str(row.get("CONCEPTO", "")).strip().upper()
        if c.startswith("IN"):
            return "COMUN INTERNO"
        if c.startswith("EX"):
            return "COMUN EXTERNO"
        return "COMUN INTERNO"

    return "COMUN INTERNO"

comunes["TIPO COSTO"] = comunes.apply(tipo_costo_hibrido, axis=1)

# Agrupar por AREA/CUENTA + TIPO COSTO (para preservar la etiqueta)
gg_agr = (
    comunes
    .groupby(["AREA/CUENTA", "TIPO COSTO"], as_index=False)["CARGOS"]
    .sum()
    .rename(columns={"CARGOS": "TOTAL_AREA"})
)

# Unir con catálogo para obtener TIPO DISTRIBUCIÓN
gg_agr = gg_agr.merge(catalogo, on="AREA/CUENTA", how="left")
if gg_agr["TIPO DISTRIBUCIÓN"].isna().any():
    faltantes = gg_agr.loc[gg_agr["TIPO DISTRIBUCIÓN"].isna(), "AREA/CUENTA"].unique().tolist()
    st.error(f"Faltan tipos de distribución en el catálogo para: {faltantes[:10]}{'...' if len(faltantes)>10 else ''}")
    st.stop()

# Expandir por sucursales según porcentajes del tipo
prorr_rows = []
for _, r in gg_agr.iterrows():
    area = r["AREA/CUENTA"]
    tipo_dist = str(r["TIPO DISTRIBUCIÓN"]).upper()
    tipo_costo = r["TIPO COSTO"]
    total = r["TOTAL_AREA"]

    if tipo_dist not in porcentajes.columns:
        st.warning(f"No hay porcentajes para el tipo '{tipo_dist}'. Se omite AREA/CUENTA: {area}")
        continue

    for suc, pct in porcentajes[tipo_dist].items():
        if pct and pct > 0:
            prorr_rows.append({
                "AREA/CUENTA": area,
                "SUCURSAL": str(suc).upper().strip(),
                "TIPO DISTRIBUCIÓN": r["TIPO DISTRIBUCIÓN"],  # conservar como está en catálogo (no upper si no quieres)
                "TIPO COSTO": tipo_costo,
                "CARGO ASIGNADO": round(total * float(pct), 2)
            })

prorr_gg = pd.DataFrame(prorr_rows)
prorr_gg = anexar_trafico_fecha(prorr_gg)

# ============ RESULTADO FINAL ============
resultado = pd.concat([directos_agr, prorr_gg], ignore_index=True)

st.subheader("📊 Resultado final (con Tráfico/Fecha por sucursal y fecha seleccionada)")
st.dataframe(resultado, use_container_width=True)

# Guardar para módulos siguientes
st.session_state["prorrateo_completo"] = resultado

# Descargar
def to_excel_bytes(df: pd.DataFrame) -> bytes:
    from io import BytesIO
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="xlsxwriter") as xw:
        df.to_excel(xw, index=False, sheet_name="Prorrateo")
    return buffer.getvalue()

st.download_button(
    "📥 Descargar prorrateo completo",
    data=to_excel_bytes(resultado),
    file_name="prorrateo_completo.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# ======================================================================================================
st.title("📘Generales (Comunes separados) e Indirectos")

# --- Obtener el prorrateo completo del Módulo 4 ---
key_candidates = ["prorrateo_completo", "prorrateo"]
prorr_key = next((k for k in key_candidates if k in st.session_state), None)

if prorr_key is None or "df_original" not in st.session_state:
    st.warning("Faltan datos. Asegúrate de haber ejecutado el Módulo 4 y tener df_original.")
    st.stop()

prorr = st.session_state[prorr_key].copy()
df_original = st.session_state["df_original"].copy()

# Normalización mínima
for df in (prorr, df_original):
    df.columns = df.columns.str.upper()

col_suc = "SUCURSAL"
col_val = "CARGO ASIGNADO"
col_tipo = "TIPO COSTO"

# 1) Comunes por sucursal (ya vienen con nombres finales)
comunes = prorr[prorr[col_tipo].isin(["COMUN INTERNO", "COMUN EXTERNO"])]

pivot_comunes = (
    comunes.pivot_table(
        index=col_suc,
        columns=col_tipo,
        values=col_val,
        aggfunc="sum",
        fill_value=0.0,
    )
    .reset_index()
)

for expected in ["COMUN INTERNO", "COMUN EXTERNO"]:
    if expected not in pivot_comunes.columns:
        pivot_comunes[expected] = 0.0

st.subheader("📌 Comunes por sucursal")
st.dataframe(pivot_comunes[[col_suc, "COMUN INTERNO", "COMUN EXTERNO"]], use_container_width=True)

# ---------- 2) Indirectos por sucursal ----------
# Suma todo lo etiquetado como COSTO INDIRECTO (incluye directos y Gasto General cuyo concepto no inicia IN/EX)
indirectos = (
    prorr[prorr[col_tipo] == "COSTO INDIRECTO"]
    .groupby(col_suc, as_index=False)[col_val]
    .sum()
    .rename(columns={col_val: "INDIRECTO"})
)

st.subheader("📌 Indirectos por sucursal")
st.dataframe(indirectos, use_container_width=True)

# ---------- 3) Consolidado: comunes separados + indirecto + total ----------
final = (
    pivot_comunes.merge(indirectos, on=col_suc, how="outer")
    .fillna(0.0)
)

final["TOTAL"] = final["COMUN INTERNO"] + final["COMUN EXTERNO"] + final["INDIRECTO"]

st.subheader("📊 Consolidado final")
st.dataframe(final[[col_suc, "COMUN INTERNO", "COMUN EXTERNO", "INDIRECTO", "TOTAL"]],
             use_container_width=True)

# ---------- 4) Exportar a Excel ----------
def exportar_excel():
    from io import BytesIO
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
        pivot_comunes[[col_suc, "COMUN INTERNO", "COMUN EXTERNO"]].to_excel(
            writer, sheet_name="Generales", index=False
        )
        indirectos.to_excel(writer, sheet_name="Indirectos", index=False)
        final[[col_suc, "COMUN INTERNO", "COMUN EXTERNO", "INDIRECTO", "TOTAL"]].to_excel(
            writer, sheet_name="Consolidado", index=False
        )
    return buffer.getvalue()

st.download_button(
    "📥 Descargar Excel (Generales/Indirectos/Consolidado)",
    data=exportar_excel(),
    file_name="generales_indirectos_consolidado.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.divider()
st.title("📅 Costo por Sucursal (Tablitas por mes)")

# =========================
# Validaciones de insumos
# =========================
if "prorrateo_completo" not in st.session_state:
    st.warning("Primero ejecuta el Módulo 4 para generar el prorrateo (prorrateo_completo).")
    st.stop()

if "df_gts" not in locals() and "df_gts" not in st.session_state:
    st.warning("Primero carga el archivo GTS (Módulo 3).")
    st.stop()

# toma df_gts ya sea del local o de session
df_gts_local = df_gts if "df_gts" in locals() else st.session_state["df_gts"]
prorr = st.session_state["prorrateo_completo"].copy()

# =========================
# Normalizaciones
# =========================
df_gts_local = df_gts_local.copy()
df_gts_local.columns = df_gts_local.columns.astype(str).str.strip().str.upper()
df_gts_local["SUCURSAL"] = df_gts_local["SUCURSAL"].astype(str).str.strip().str.upper()

prorr.columns = prorr.columns.astype(str).str.strip().str.upper()
prorr["SUCURSAL"] = prorr["SUCURSAL"].astype(str).str.strip().str.upper()

# =========================
# Helper: detectar columnas en GTS
# =========================
def _find_col(df, candidates):
    cols = list(df.columns)
    for c in candidates:
        if c in cols:
            return c
    # fallback: búsqueda parcial
    for c in cols:
        for cand in candidates:
            if cand in c:
                return c
    return None

COL_FACT = _find_col(df_gts_local, ["FACTURACION DLLS", "FACTURACIÓN DLLS", "FACTURACION", "FACTURACIÓN"])
COL_MC   = _find_col(df_gts_local, ["MC", "M.C.", "MARGEN", "MARGEN CONTRIBUCION", "MARGEN DE CONTRIBUCION"])

if not COL_FACT or not COL_MC:
    st.error(
        "No pude detectar columnas en GTS.\n"
        f"Encontré: {list(df_gts_local.columns)}\n\n"
        "Asegúrate de tener algo como 'FACTURACION DLLS' y 'MC'."
    )
    st.stop()

# =========================
# Catálogo para distribución (AREA/CUENTA -> TIPO DISTRIBUCIÓN)
# =========================
cat_resp = supabase.table("catalogo_distribucion").select("*").execute()
catalogo = pd.DataFrame(cat_resp.data)
if catalogo.empty:
    st.error("No hay datos en 'catalogo_distribucion' para traer el método de distribución.")
    st.stop()

catalogo = catalogo.rename(columns={
    "area_cuenta": "AREA/CUENTA",
    "tipo_distribucion": "TIPO DISTRIBUCIÓN"
})
catalogo["AREA/CUENTA"] = catalogo["AREA/CUENTA"].astype(str).str.strip().str.upper()


# =========================
# Función principal: genera 3 tablitas
# =========================
def generar_tablitas_mes_sucursal(sucursal: str):
    suc = str(sucursal).strip().upper()

    # ---- 1) Datos principales desde GTS ----
    row_gts = df_gts_local[df_gts_local["SUCURSAL"].eq(suc)]
    if row_gts.empty:
        return None, None, None, f"No existe la sucursal '{suc}' en el archivo GTS."

    facturacion = float(row_gts.iloc[0][COL_FACT] or 0)
    mc = float(row_gts.iloc[0][COL_MC] or 0)

    # costos directos = facturación - MC
    costos_directos = facturacion - mc
    utilidad = mc

    # ---- 2) GASTOS INDIRECTOS: SOLO COMUN INTERNO (prorrateo ya asignado por sucursal) ----
    gi = prorr[(prorr["SUCURSAL"].eq(suc)) & (prorr["TIPO COSTO"].eq("COSTO INDIRECTO"))].copy()

    tabla_gi = (
        gi.groupby("AREA/CUENTA", as_index=False)["CARGO ASIGNADO"]
          .sum()
          .rename(columns={"AREA/CUENTA": "GASTOS INDIRECTOS", "CARGO ASIGNADO": "IMPORTE"})
          .sort_values("IMPORTE", ascending=False)
          .reset_index(drop=True)
    )
    tabla_gi["%"] = np.where(facturacion != 0, tabla_gi["IMPORTE"] / facturacion, 0.0)

    total_ci = float(tabla_gi["IMPORTE"].sum())
    pct_ci = (total_ci / facturacion) if facturacion != 0 else 0.0

    # ---- 3) AREA-TIPO GASTO: SOLO COMUN EXTERNO + método distribución desde catálogo ----
    ge = prorr[(prorr["SUCURSAL"].eq(suc)) & (prorr["TIPO COSTO"].isin(["COMUN EXTERNO", "COMUN INTERNO"]))].copy()
    tabla_ge = (
        ge.groupby("AREA/CUENTA", as_index=False)["CARGO ASIGNADO"]
          .sum()
          .rename(columns={"AREA/CUENTA": "AREA-TIPO GASTO", "CARGO ASIGNADO": "IMPORTE"})
    )

    # agregar distribución desde catálogo
    tabla_ge["AREA_KEY"] = tabla_ge["AREA-TIPO GASTO"].astype(str).str.strip().str.upper()
    tabla_ge = tabla_ge.merge(
        catalogo[["AREA/CUENTA", "TIPO DISTRIBUCIÓN"]].rename(columns={"AREA/CUENTA": "AREA_KEY"}),
        on="AREA_KEY",
        how="left"
    ).drop(columns=["AREA_KEY"])

    tabla_ge["%"] = np.where(facturacion != 0, tabla_ge["IMPORTE"] / facturacion, 0.0)
    tabla_ge = tabla_ge.sort_values("IMPORTE", ascending=False).reset_index(drop=True)

    total_gn = float(tabla_ge["IMPORTE"].sum())
    pct_gn = (total_gn / facturacion) if facturacion != 0 else 0.0

    # ---- 4) Tablita superior ----
    pct_ut_bruta = (utilidad / facturacion) if facturacion != 0 else 0.0
    ut_per = utilidad - total_ci - total_gn
    pct_ut_per = (ut_per / facturacion) if facturacion != 0 else 0.0

    tabla_top = pd.DataFrame([{
        "Sucursal": suc,
        "Facturación": facturacion,
        "Costos Directos": costos_directos,
        "Utilidad": utilidad,
        "% Ut Bruta": pct_ut_bruta,
        "Costos Indirectos": total_ci,
        "% CI": pct_ci,
        "Gastos Generales": total_gn,
        "% GN": pct_gn,
        "UT/PER": ut_per,
        "%UT/PER": pct_ut_per
    }])

    return tabla_top, tabla_gi, tabla_ge, None


# =========================
# UI
# =========================
sucursales_disponibles = sorted(df_gts_local["SUCURSAL"].dropna().unique().tolist())
sucursal_sel = st.selectbox("Selecciona sucursal", sucursales_disponibles)

tabla_top, tabla_gi, tabla_ge, err = generar_tablitas_mes_sucursal(sucursal_sel)

if err:
    st.error(err)
    st.stop()

# ---- Mostrar tablita superior ----
st.subheader("🧾 Resumen superior")
st.dataframe(tabla_top, use_container_width=True)

# ---- Mostrar gastos indirectos (COMUN INTERNO) ----
st.subheader("📌 GASTOS INDIRECTOS (Costo indirecto)")
st.dataframe(tabla_gi, use_container_width=True)

# ---- Mostrar area-tipo gasto (COMUN EXTERNO) ----
st.subheader("📌 AREA-TIPO GASTO (Común interno + Común externo)")
st.dataframe(tabla_ge, use_container_width=True)


# =========================
# Export a Excel (1 hoja por sucursal)
# =========================
def exportar_costo_sucursal_excel(tabla_top, tabla_gi, tabla_ge, nombre_sucursal):
    from io import BytesIO
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
        sheet = str(nombre_sucursal)[:31]  # límite Excel
        # posiciones estilo "layout" (sin formato heavy)
        tabla_top.to_excel(writer, sheet_name=sheet, index=False, startrow=0, startcol=0)
        tabla_gi.to_excel(writer, sheet_name=sheet, index=False, startrow=4, startcol=0)
        tabla_ge.to_excel(writer, sheet_name=sheet, index=False, startrow=4, startcol=6)

    return buffer.getvalue()

st.download_button(
    "📥 Descargar Excel de esta sucursal",
    data=exportar_costo_sucursal_excel(tabla_top, tabla_gi, tabla_ge, sucursal_sel),
    file_name=f"costo_por_sucursal_{sucursal_sel}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# =========================
# ESTA PARTE CONSOLIDA TODOS LOS RESUMENES MENSUALES
# =========================
st.divider()
st.title("📚 Consolidar histórico (archivo con tablitas por mes)")

archivo_hist = st.file_uploader(
    "Sube el Excel histórico (una hoja por sucursal, con bloques mensuales tipo 'ENERO 2025')",
    type=["xlsx"],
    key="historico_excel"
)

MESES_ES = [
    "ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
    "JULIO","AGOSTO","SEPTIEMBRE","SETIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"
]

def _es_titulo_mes(v) -> bool:
    if not isinstance(v, str):
        return False
    s = v.strip().upper()
    return any(m in s for m in MESES_ES) and re.search(r"\b20\d{2}\b", s)

def _extraer_mes(v) -> str:
    # "ENERO 2025" -> "ENERO"
    s = str(v).strip().upper()
    for m in MESES_ES:
        if m in s:
            return m
    return s

def _orden_mes(m):
    d = {m:i+1 for i,m in enumerate(["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO","JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"])}
    return d.get(str(m).strip().upper(), 999)

# Encabezados esperados en la tablita superior (los de tu screenshot)
HEADS = ["FACTURACIÓN","COSTOS DIRECTOS","UTILIDAD","% UT BRUTA","COSTOS INDIRECTOS","% CI",
         "GASTOS GENERALES","% GN","UT/PER","%UT/PER"]

def _normaliza_header(h):
    s = str(h).strip().upper()
    s = s.replace("FACTURACION","FACTURACIÓN")  # por si viene sin acento
    s = s.replace("%UT/PER","%UT/PER")
    return s

def parse_sheet(ws, sheet_name: str) -> pd.DataFrame:
    """
    Busca todos los títulos tipo 'ENERO 2025' en la hoja.
    Para cada título encontrado:
      - fila+1: headers
      - fila+2: valores
    Construye dataframe consolidado.
    """
    rows = []

    max_r = ws.max_row
    max_c = ws.max_column

    # escaneo completo (tu formato tiene títulos en la fila 1, pero esto lo hace robusto)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if not _es_titulo_mes(v):
                continue

            mes = _extraer_mes(v)

            # headers en r+1, valores en r+2
            r_head = r + 1
            r_val  = r + 2
            if r_val > max_r:
                continue

            # tomamos 10 columnas desde c (como en tu layout)
            headers = [ws.cell(r_head, c+i).value for i in range(0, 10)]
            values  = [ws.cell(r_val,  c+i).value for i in range(0, 10)]

            headers = [_normaliza_header(h) for h in headers]

            # Validación mínima: que existan columnas clave
            if "FACTURACIÓN" not in headers and "FACTURACION" not in headers:
                continue
            if "UT/PER" not in headers:
                continue

            # arma dict
            d = {"Mes": mes, "Sucursal": str(sheet_name).strip().upper()}
            for h, val in zip(headers, values):
                d[h] = val

            # renombres a tu formato final
            # (así queda igual que tu tabla consolidada)
            out = {
                "Mes": d.get("Mes"),
                "Sucursal": d.get("Sucursal"),
                "Facturación": float(d.get("FACTURACIÓN") or 0),
                "Costos Dire": float(d.get("COSTOS DIRECTOS") or 0),
                "Utilidad": float(d.get("UTILIDAD") or 0),
                "% Ut Bruta": float(d.get("% UT BRUTA") or 0),
                "Costos Indirectos": float(d.get("COSTOS INDIRECTOS") or 0),
                "% CI": float(d.get("% CI") or 0),
                "Gastos Generales": float(d.get("GASTOS GENERALES") or 0),
                "% GN": float(d.get("% GN") or 0),
                "UT/PER": float(d.get("UT/PER") or 0),
                "%UT/PER": float(d.get("%UT/PER") or 0),
            }

            rows.append(out)

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)

    # quitar duplicados si existen (por si re-scan encuentra el mismo bloque)
    df = df.drop_duplicates(subset=["Mes","Sucursal"], keep="last")

    # ordenar por mes
    df["__m"] = df["Mes"].apply(_orden_mes)
    df = df.sort_values(["Sucursal","__m"]).drop(columns="__m").reset_index(drop=True)

    # orden final
    cols = ["Mes","Sucursal","Facturación","Costos Dire","Utilidad","% Ut Bruta",
            "Costos Indirectos","% CI","Gastos Generales","% GN","UT/PER","%UT/PER"]
    return df[cols]

def exportar_consolidado_por_sucursal(dfs_por_sucursal: dict) -> bytes:
    from io import BytesIO
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
        for suc, df in dfs_por_sucursal.items():
            if df.empty:
                continue
            sheet = str(suc)[:31]
            df.to_excel(writer, sheet_name=sheet, index=False)
    return buffer.getvalue()

if archivo_hist is not None:
    wb = openpyxl.load_workbook(archivo_hist, data_only=True)

    dfs = {}
    for sh in wb.sheetnames:
        # opcional: saltar CONSOLIDADO si no lo quieres
        # if sh.strip().upper() == "CONSOLIDADO":
        #     continue

        ws = wb[sh]
        df_sh = parse_sheet(ws, sh)
        if not df_sh.empty:
            dfs[sh] = df_sh

    if not dfs:
        st.error("No encontré bloques mensuales tipo 'ENERO 2025' en las hojas. Revisa el formato del archivo.")
    else:
        st.success(f"Listo: detecté consolidado en {len(dfs)} hojas.")

        sucursales = sorted(dfs.keys())
        suc_sel = st.selectbox("Ver consolidado de sucursal", sucursales, key="ver_suc_hist")
        st.dataframe(dfs[suc_sel], use_container_width=True)

        st.download_button(
            "📥 Descargar consolidado (1 hoja por sucursal)",
            data=exportar_consolidado_por_sucursal(dfs),
            file_name="consolidado_historico_por_sucursal.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
