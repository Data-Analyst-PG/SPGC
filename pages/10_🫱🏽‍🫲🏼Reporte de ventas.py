from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ============================================================
# Configuración del módulo
# ============================================================
ETAPA_COLUMNS = ["1.1", "1.2", "1.3", "2.1", "2.2", "2.3", "2.4", "2.5", "2.6", "3.1", "3.2", "4.1", "4.2"]
RESUMEN_COLUMNS = ["Id", "Cliente", "Tipo", *ETAPA_COLUMNS, "Vendedor"]
DETALLE_COLUMNS = [
    "Id",
    "Cliente",
    "Fecha",
    "Etapa",
    "Comentarios",
    "Fecha siguiente Evento",
    "Comentarios Siguiente Evento",
    "Medio de Contacto",
    "Atiende Cita",
    "Registro",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_GRAY = Side(style="thin", color="D9D9D9")
BORDER = Border(bottom=THIN_GRAY)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


# ============================================================
# Modelos
# ============================================================
@dataclass
class ParseResult:
    resumen: pd.DataFrame
    detalle: pd.DataFrame
    periodo_texto: str | None = None


# ============================================================
# Utilidades de parsing
# ============================================================
def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _as_datetime(value: Any) -> Any:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return value
    return parsed.to_pydatetime()


def _is_page_footer(row_values: list[Any]) -> bool:
    joined = " | ".join(_normalize_text(v) for v in row_values if v not in (None, ""))
    joined_lower = joined.lower()
    return "página de" in joined_lower or "mostrando" in joined_lower


def _extract_periodo(ws) -> str | None:
    patron = re.compile(r"Seguimiento\s+Periodo\s+de\s*:\s*(.+)", re.IGNORECASE)
    for row in ws.iter_rows(values_only=True):
        for value in row:
            text = _normalize_text(value)
            if not text:
                continue
            match = patron.search(text)
            if match:
                return match.group(1).strip()
    return None


def _find_header_row(ws) -> int:
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [_normalize_text(v) for v in row]
        if "Id" in values and "Cliente" in values and "Vendedor" in values:
            return idx
    raise ValueError("No encontré la fila de encabezados principales (Id / Cliente / Vendedor).")


def _build_column_map(ws, header_row: int) -> dict[str, int]:
    colmap: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = _normalize_text(ws.cell(header_row, col).value)
        if value:
            colmap[value] = col
    return colmap


def _extract_stage_columns(ws, header_row: int) -> dict[str, int]:
    stage_map: dict[str, int] = {}
    scan_row = header_row + 1
    for col in range(1, ws.max_column + 1):
        value = ws.cell(scan_row, col).value
        text = _normalize_text(value)
        if text in ETAPA_COLUMNS:
            stage_map[text] = col
    missing = [c for c in ETAPA_COLUMNS if c not in stage_map]
    if missing:
        raise ValueError(f"Faltan columnas de etapas en la descarga: {', '.join(missing)}")
    return stage_map


def _is_summary_row(ws, row_idx: int, id_col: int, vendedor_col: int) -> bool:
    id_val = ws.cell(row_idx, id_col).value
    vendedor_val = ws.cell(row_idx, vendedor_col).value
    if id_val in (None, "") or vendedor_val in (None, ""):
        return False
    next_cell = _normalize_text(ws.cell(row_idx + 1, id_col).value)
    return next_cell.lower().startswith("seguimiento")


def _safe_int(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(value))
    except Exception:
        return 0


def parse_seguimiento_workbook(file_bytes: bytes) -> ParseResult:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if len(wb.sheetnames) != 1:
        raise ValueError("El archivo debe contener una sola hoja con la descarga pegada desde el sistema.")

    ws = wb[wb.sheetnames[0]]
    periodo = _extract_periodo(ws)
    header_row = _find_header_row(ws)
    colmap = _build_column_map(ws, header_row)
    stage_map = _extract_stage_columns(ws, header_row)

    required_headers = ["Id", "Cliente", "Tipo", "Vendedor"]
    missing_headers = [h for h in required_headers if h not in colmap]
    if missing_headers:
        raise ValueError(f"Faltan columnas requeridas en la descarga: {', '.join(missing_headers)}")

    resumen_rows: list[dict[str, Any]] = []
    detalle_rows: list[dict[str, Any]] = []

    current_cliente: str | None = None
    current_id: Any = None

    row_idx = header_row + 1
    while row_idx <= ws.max_row:
        row_values = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]

        if _is_page_footer(row_values):
            row_idx += 1
            continue

        if _is_summary_row(ws, row_idx, colmap["Id"], colmap["Vendedor"]):
            current_id = ws.cell(row_idx, colmap["Id"]).value
            current_cliente = ws.cell(row_idx, colmap["Cliente"]).value

            resumen_item: dict[str, Any] = {
                "Id": current_id,
                "Cliente": current_cliente,
                "Tipo": ws.cell(row_idx, colmap["Tipo"]).value,
                "Vendedor": ws.cell(row_idx, colmap["Vendedor"]).value,
            }
            for etapa in ETAPA_COLUMNS:
                resumen_item[etapa] = _safe_int(ws.cell(row_idx, stage_map[etapa]).value)
            resumen_rows.append(resumen_item)
            row_idx += 1
            continue

        first_value = _normalize_text(ws.cell(row_idx, colmap["Id"]).value)
        if first_value == "Fecha":
            row_idx += 1
            continue

        fecha = ws.cell(row_idx, colmap["Id"]).value
        etapa = ws.cell(row_idx, 8).value  # Columna H en la descarga original
        registro = ws.cell(row_idx, 21).value  # Columna U en la descarga original

        is_detail_row = any(
            x not in (None, "")
            for x in [fecha, etapa, ws.cell(row_idx, 17).value, ws.cell(row_idx, 16).value, registro]
        )

        if is_detail_row and current_id not in (None, ""):
            detalle_rows.append(
                {
                    "Id": current_id,
                    "Cliente": current_cliente,
                    "Fecha": _as_datetime(fecha),
                    "Etapa": etapa,
                    "Comentarios": ws.cell(row_idx, 16).value,
                    "Fecha siguiente Evento": _as_datetime(ws.cell(row_idx, 17).value),
                    "Comentarios Siguiente Evento": ws.cell(row_idx, 18).value,
                    "Medio de Contacto": ws.cell(row_idx, 19).value,
                    "Atiende Cita": ws.cell(row_idx, 20).value,
                    "Registro": registro,
                }
            )

        row_idx += 1

    resumen_df = pd.DataFrame(resumen_rows, columns=RESUMEN_COLUMNS)
    detalle_df = pd.DataFrame(detalle_rows, columns=DETALLE_COLUMNS)

    if resumen_df.empty:
        raise ValueError("No se pudo construir la hoja Resumen. Revisa que la descarga venga en el formato esperado del sistema.")

    if not detalle_df.empty:
        detalle_df = detalle_df.sort_values(["Id", "Fecha"], ascending=[True, False], na_position="last").reset_index(drop=True)
    resumen_df = resumen_df.sort_values(["Vendedor", "Cliente"], ascending=[True, True], na_position="last").reset_index(drop=True)

    return ParseResult(resumen=resumen_df, detalle=detalle_df, periodo_texto=periodo)


# ============================================================
# Escritura del archivo de salida
# ============================================================
def _autosize(ws) -> None:
    max_widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            max_widths[cell.column] = max(max_widths.get(cell.column, 0), len(value))
    for col_idx, width in max_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 40)


def _style_table_sheet(ws, title: str, freeze_cell: str = "A2") -> None:
    ws.freeze_panes = freeze_cell
    ws.sheet_view.showGridLines = False
    ws.insert_rows(1)
    ws["A1"] = title
    ws["A1"].font = Font(size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = LEFT
    max_col = ws.max_column
    if max_col > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

    for cell in ws[2]:
        cell.fill = SUBHEADER_FILL
        cell.font = BOLD_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = LEFT
            cell.border = Border(bottom=THIN_GRAY)

    _autosize(ws)


def _format_dates(ws, date_columns: list[str]) -> None:
    headers = {cell.value: idx for idx, cell in enumerate(ws[2], start=1)}
    for name in date_columns:
        col = headers.get(name)
        if not col:
            continue
        for row in range(3, ws.max_row + 1):
            ws.cell(row, col).number_format = "dd/mm/yyyy"


def _add_excel_table(ws, table_name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def build_output_workbook(parse_result: ParseResult) -> bytes:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        parse_result.resumen.to_excel(writer, index=False, sheet_name="Resumen")
        parse_result.detalle.to_excel(writer, index=False, sheet_name="Detalle")

        wb = writer.book
        ws_resumen = wb["Resumen"]
        ws_detalle = wb["Detalle"]

        resumen_title = "Resumen"
        detalle_title = "Detalle"
        if parse_result.periodo_texto:
            resumen_title = f"Resumen | Periodo: {parse_result.periodo_texto}"
            detalle_title = f"Detalle | Periodo: {parse_result.periodo_texto}"

        _style_table_sheet(ws_resumen, resumen_title)
        _style_table_sheet(ws_detalle, detalle_title)
        _format_dates(ws_detalle, ["Fecha", "Fecha siguiente Evento"])
        _add_excel_table(ws_resumen, "tblResumenSeguimiento")
        _add_excel_table(ws_detalle, "tblDetalleSeguimiento")

    output.seek(0)
    return output.getvalue()


# ============================================================
# Streamlit UI del módulo
# ============================================================
def render_modulo_seguimiento_sac_ventas() -> None:
    st.subheader("Seguimiento SAC Ventas")
    st.caption(
        "Sube un Excel con una sola hoja que contenga la descarga pegada desde el sistema. "
        "El módulo generará un archivo con las hojas Resumen y Detalle."
    )

    uploaded = st.file_uploader(
        "Archivo Excel de entrada",
        type=["xlsx", "xlsm"],
        help="Debe contener únicamente una hoja, sin importar el nombre de la hoja ni del archivo.",
    )

    if not uploaded:
        return

    file_bytes = uploaded.getvalue()

    try:
        result = parse_seguimiento_workbook(file_bytes)
        out_bytes = build_output_workbook(result)
    except Exception as exc:
        st.error(f"No fue posible procesar el archivo: {exc}")
        with st.expander("Ver recomendaciones"):
            st.markdown(
                """
                - Asegúrate de que el archivo tenga **una sola hoja**.
                - Pega la descarga del sistema completa, sin eliminar filas intermedias.
                - No conviertas el contenido a tabla antes de subirlo.
                - Conserva las columnas y bloques tal como los entrega el sistema.
                """
            )
        return

    c1, c2 = st.columns(2)
    with c1:
        st.metric("Registros en resumen", len(result.resumen))
    with c2:
        st.metric("Registros en detalle", len(result.detalle))

    if result.periodo_texto:
        st.info(f"Periodo detectado: {result.periodo_texto}")

    with st.expander("Vista previa - Resumen", expanded=True):
        st.dataframe(result.resumen, use_container_width=True, hide_index=True)

    with st.expander("Vista previa - Detalle"):
        st.dataframe(result.detalle, use_container_width=True, hide_index=True)

    output_name = f"reporte_seguimiento_sac_ventas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    st.download_button(
        "Descargar reporte Excel",
        data=out_bytes,
        file_name=output_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


# ============================================================
# Ejecución standalone opcional
# ============================================================
if __name__ == "__main__":
    st.set_page_config(page_title="Seguimiento SAC Ventas", layout="wide")
    render_modulo_seguimiento_sac_ventas()
